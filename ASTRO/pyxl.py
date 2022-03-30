from openpyxl import Workbook

# 엑셀파일 쓰기
write_wb = Workbook()

# 이름이 있는 시트를 생성
write_ws = write_wb.create_sheet('태양의 고도 및 방위각 변화')

# Sheet1에다 입력
write_ws = write_wb.active

#행 단위로 추가
write_ws.append(['년도-월-일-시간(시)', '방위각(도 분 초)', '고도(도 분 초)', '적경(시 분 초)', '적위(도 분 초)'])
write_ws.append(['2021-06-04-00시',	'351 55 32.49',	'-29 38 26.69',	'04 47 21.29',	'22 23 07.86'])

#셀 단위로 추가
write_wb.save("/Users/osanghyun/PycharmProjects/DBProject2/ASTRO/EXCEL_FILE/예제_2021.xlsx")

# 새 파일
write_wb = Workbook()

# 이름이 있는 시트를 생성
write_ws = write_wb.create_sheet('태양의 고도 및 방위각 변화')

# Sheet1에다 입력
write_ws = write_wb.active

#행 단위로 추가
write_ws.append(['년도-ㅋㅋ-일-시간(시)', '방위각(도 분 초)', '고도(도 분 초)', '적경(시 분 초)', '적위(도 분 초)'])
write_ws.append(['2021-ㅋㅋ-04-00시',	'351 55 32.49',	'-29 38 26.69',	'04 47 21.29',	'22 23 07.86'])


#셀 단위로 추가
write_wb.save("/Users/osanghyun/PycharmProjects/DBProject2/ASTRO/EXCEL_FILE/예제_new.xlsx")