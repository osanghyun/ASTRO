from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
import selenium
import sys
import calender
from calender import *
import re
import time
from openpyxl import Workbook


def element_presence_wait(driver, by_type, locator: str):
    """
    데이터가 존재하는지 체크.
    :param by_type:
    :param locator:
    :return: 존재 유무
    """
    try:
        WebDriverWait(driver, 3, poll_frequency=0.01).until(EC.presence_of_element_located((by_type, locator)))
        print(f"is_element_presence 함수 호출 ({by_type}, {locator}) 결과 : 존재")

    except selenium.common.exceptions.TimeoutException:
        print(f"is_element_presence 함수 호출 ({by_type}, {locator}) 결과 : 비존재")
        print("오류")
        sys.exit(-1)


driver = webdriver.Chrome('./chromedriver')

# 천문우주지식정보 접속.
driver.get("https://astro.kasi.re.kr/index")
assert "천문우주지식정보" in driver.title

# 생활천문관 접속.
driver.find_element(by=By.LINK_TEXT, value="생활천문관").click()
assert "생활천문관 | 천문우주지식정보" in driver.title

# 생활천문관 접속.
element_presence_wait(driver, By.LINK_TEXT, "태양 고도/방위각계산")
driver.find_element(by=By.LINK_TEXT, value="태양 고도/방위각계산").click()
assert "태양 고도/방위각계산 | 생활천문관 | 천문우주지식정보" in driver.title

# 위치입력
# 울산광역시 남구 용잠로 623
location_button = driver.find_element(by=By.ID, value="addrBtn")
location_button.click()

main = driver.window_handles

if len(main) == 2:
    driver.switch_to.window(main[1])

    if '안전하지 않은 양식' in driver.title:
        proceed_button = driver.find_element(by=By.ID, value="proceed-button")
        proceed_button.click()

    text_input = driver.find_element(by=By.CLASS_NAME, value="popSearchInput")
    text_input.send_keys("울산광역시 남구 용잠로 623")

    submit_button = driver.find_element(by=By.CSS_SELECTOR,
                                        value="#serarchContentBox > div.search-wrap > fieldset > span "
                                              "> input[type=button]:nth-child(2)")
    submit_button.click()

    result_elem = driver.find_element(by=By.CSS_SELECTOR, value="#roadAddrDiv1 > b")
    result_elem.click()

    addr_input_button = driver.find_element(by=By.CSS_SELECTOR, value="#resultData > div > a")
    addr_input_button.click()

    driver.switch_to.window(main[0])
    search_button = driver.find_element(by=By.CSS_SELECTOR, value="#del_btn")
    search_button.click()
else:
    print("addr error")
    sys.exit(-1)

# 2019년 1월 1일 찾기
calender_elem = driver.find_element(by=By.CSS_SELECTOR, value="#dp > a")
calender_elem.click()

year_month_day = datetime.date(2019, 1, 1)

while True:
    month_year_elem = driver.find_element(by=By.CLASS_NAME, value="switch")

    info = month_year_elem.text

    month_pattern = re.compile(r'[\d]+월')
    year_pattern = re.compile(r'[\d]{4}')

    month_search_data = month_pattern.search(info)
    year_search_data = year_pattern.search(info)

    month = int(month_search_data[0][:-1])
    year = int(year_search_data[0])

    if year_month_day.year == year and year_month_day.month == month:
        print("match success")
        break
    else:
        prev_elem = driver.find_element(by=By.CLASS_NAME, value="prev")
        prev_elem.click()

# 일 찾기
calender_body = driver.find_element(by=By.CSS_SELECTOR,
                                    value="body > div.datepicker.dropdown-menu > div.datepicker-days > table > tbody")
trs = calender_body.find_elements(by=By.TAG_NAME, value="tr")

for tr in trs:
    status = 0
    tds = tr.find_elements(by=By.TAG_NAME, value="td")

    for td in tds:
        class_name = td.get_attribute(name="class")
        if 'old' not in class_name and 'new' not in class_name and int(td.text) == year_month_day.day:
            td.click()
            status = 1
            break

    if status == 1:
        break

# 엑셀파일 초기화
excel_year = 2018
write_wb = Workbook()
write_ws = write_wb.create_sheet('태양의 고도 및 방위각 변화')
write_ws = write_wb.active

# 테이블 내용 크롤링
crawl_year = str(year_month_day.year)
crawl_month = str(year_month_day.month)
if len(crawl_month) == 1:
    crawl_month = '0' + crawl_month
crawl_day = str(year_month_day.day)
if len(crawl_day) == 1:
    crawl_day = '0' + crawl_day

element_presence_wait(driver, By.CSS_SELECTOR, "#sun-height-table > table > tbody")
table_body = driver.find_element(by=By.CSS_SELECTOR, value="#sun-height-table > table > tbody")

trs = table_body.find_elements(by=By.TAG_NAME, value="tr")
for tr in trs:
    list_data = []
    tds = tr.find_elements(by=By.TAG_NAME, value="td")
    list_data.append(f'{crawl_year}-{crawl_month}-{crawl_day}-{tds[0].text}시')
    list_data.append(tds[1].text)
    list_data.append(tds[2].text)
    list_data.append(tds[3].text)
    list_data.append(tds[4].text)

    write_ws.append(list_data)


# 2019-01-02 부터 2021-06-03 까지
year_month_day_list = calender.get_yeardatetime_list()

for year_month_day in year_month_day_list:
    print(year_month_day)
    if year_month_day.year != excel_year:
        write_wb.save(f'/Users/osanghyun/PycharmProjects/DBProject2/ASTRO/EXCEL_FILE/태양의_고도_및_방위각_변화{excel_year}.xlsx')
        write_wb = Workbook()
        write_ws = write_wb.create_sheet('태양의 고도 및 방위각 변화')
        write_ws = write_wb.active
        excel_year = year_month_day.year

    # 캘린더 아이콘 클릭
    calender_elem = driver.find_element(by=By.CSS_SELECTOR, value="#dp > a")
    calender_elem.click()

    # 년 월 찾기
    while True:
        month_year_elem = driver.find_element(by=By.CLASS_NAME, value="switch")

        info = month_year_elem.text

        month_pattern = re.compile(r'[\d]+월')
        year_pattern = re.compile(r'[\d]{4}')

        month_search_data = month_pattern.search(info)
        year_search_data = year_pattern.search(info)

        month = int(month_search_data[0][:-1])
        year = int(year_search_data[0])

        if year_month_day.year == year and year_month_day.month == month:
            print("match success")
            break
        else:
            next_elem = driver.find_element(by=By.CLASS_NAME, value="next")
            next_elem.click()

    # 일 찾기
    calender_body = driver.find_element(by=By.CSS_SELECTOR,
                                        value="body > div.datepicker.dropdown-menu > div.datepicker-days > table > tbody")
    trs = calender_body.find_elements(by=By.TAG_NAME, value="tr")

    for tr in trs:
        status = 0
        tds = tr.find_elements(by=By.CLASS_NAME, value="day ")

        for td in tds:
            if int(td.text) == year_month_day.day:
                td.click()
                status = 1
                break

        if status == 1:
            break

    search_button = driver.find_element(by=By.CSS_SELECTOR, value="#del_btn")
    search_button.click()

    # 테이블 내용 크롤링
    crawl_year = str(year_month_day.year)
    crawl_month = str(year_month_day.month)
    if len(crawl_month) == 1:
        crawl_month = '0' + crawl_month
    crawl_day = str(year_month_day.day)
    if len(crawl_day) == 1:
        crawl_day = '0' + crawl_day

    element_presence_wait(driver, By.CSS_SELECTOR, "#sun-height-table > table > tbody")
    table_body = driver.find_element(by=By.CSS_SELECTOR, value="#sun-height-table > table > tbody")

    trs = table_body.find_elements(by=By.TAG_NAME, value="tr")
    for tr in trs:
        list_data = []
        tds = tr.find_elements(by=By.TAG_NAME, value="td")
        list_data.append(f'{crawl_year}-{crawl_month}-{crawl_day}-{tds[0].text}시')
        list_data.append(tds[1].text)
        list_data.append(tds[2].text)
        list_data.append(tds[3].text)
        list_data.append(tds[4].text)

        write_ws.append(list_data)

write_wb.save(f'/Users/osanghyun/PycharmProjects/DBProject2/ASTRO/EXCEL_FILE/태양의_고도_및_방위각_변화{excel_year}.xlsx')